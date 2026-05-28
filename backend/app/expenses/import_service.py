"""
File Import Service — UPI / Bank Statement Import
===================================================
Parses CSV, PDF, and Excel (.xlsx) exports from major Indian payment apps and banks:
  - PhonePe  (CSV from app  •  PDF statement)
  - Google Pay (CSV from myaccount.google.com)
  - Paytm    (CSV from Passbook  •  PDF from Paytm Bank  •  Paytm UPI Statement PDF)
  - HDFC Bank (CSV  •  PDF statement)
  - SBI       (CSV  •  PDF statement)
  - Generic   (any CSV/PDF/Excel with Date, Description, Debit/Credit columns)

Entry point:
    parse_file(content: bytes, filename: str, user_categories) → ImportPreview

Flow:
  1. Detect file type from extension / magic bytes (.pdf vs .csv vs .xlsx)
  2. For PDF → detect format, use specialized parser or table extraction
  3. For XLSX → openpyxl → reconstruct rows
  4. For CSV → standard csv.DictReader
  5. Detect bank format from column headers
  6. Parse rows → list[ParsedTransaction]
  7. Match each description → best-fit category via keyword table
  8. Return ImportPreview (user reviews → confirms → expenses created)

Why pdfplumber for PDF?
  pdfplumber uses pdfminer under the hood but adds table-detection heuristics
  and word-position extraction. We use extract_tables() for bordered tables
  (HDFC, SBI, standard banks) and extract_words() for card-layout PDFs
  like Paytm UPI Statement which has no visible table borders.

Why keyword matching instead of AI?
  Fast, offline, zero API cost, works without a network.
  Handles ~80 % of common merchant names correctly.
  The remaining 20 % can be corrected on the preview screen before confirming.
"""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

import structlog

logger = structlog.get_logger(__name__)


# =============================================================================
# DATA MODELS
# =============================================================================


@dataclass
class ParsedTransaction:
    """One parsed transaction row, ready for preview."""

    row_index: int
    date: str  # ISO format: "2026-05-01"
    description: str  # Cleaned merchant / narration text
    amount: Decimal  # Always positive
    transaction_type: str  # "DEBIT" or "CREDIT"
    reference: str  # UPI ref / transaction ID
    suggested_category_name: str
    suggested_category_id: str  # UUID string (or "" if no match)
    raw: str  # Original row for debugging


@dataclass
class ImportPreview:
    """Full preview returned to the UI before the user confirms import."""

    format_detected: str  # "phonepe" | "gpay" | "paytm" | "hdfc" | "sbi" | "generic"
    file_type: str  # "csv" | "pdf" | "xlsx"
    total_rows: int
    debit_count: int
    credit_count: int
    skipped_count: int
    transactions: list[ParsedTransaction]


# =============================================================================
# CATEGORY KEYWORD TABLE
# =============================================================================

# Lower-case keywords — checked with `in` on the lower-cased description.
# Order matters: first match wins.
CATEGORY_KEYWORDS: dict[str, list[str]] = {
    "Food & Dining": [
        "swiggy",
        "zomato",
        "ubereats",
        "dunzo",
        "blinkit food",
        "food",
        "tiffin",
        "biryani",
        "maggi",
    ],
    "Restaurants": [
        "mcdonalds",
        "kfc",
        "dominos",
        "pizza hut",
        "subway",
        "burger king",
        "restaurant",
        "cafe",
        "dhaba",
        "chai",
        "coffee",
        "starbucks",
        "barista",
        "ccd",
    ],
    "Groceries": [
        "bigbasket",
        "grofers",
        "blinkit",
        "zepto",
        "dmart",
        "more super",
        "reliance fresh",
        "jiomart",
        "supermarket",
        "grocery",
        "kirana",
        "vegetable",
        "fruit",
        "milk",
    ],
    "Shopping": [
        "amazon",
        "flipkart",
        "myntra",
        "ajio",
        "nykaa",
        "meesho",
        "snapdeal",
        "shopclues",
        "shopping",
        "lenskart",
        "pepperfry",
        "ikea",
    ],
    "Transportation": [
        "uber",
        "ola",
        "rapido",
        "namma yatri",
        "metro",
        "petrol",
        "fuel",
        "hp petrol",
        "bpcl",
        "iocl",
        "indianoil",
        "auto rickshaw",
        "bus",
        "ksrtc",
        "bmtc",
    ],
    "Entertainment": [
        "netflix",
        "amazon prime",
        "hotstar",
        "disney+",
        "zee5",
        "sonyliv",
        "bookmyshow",
        "pvr",
        "inox",
        "spotify",
        "youtube premium",
        "gaana",
        "jiosavaan",
    ],
    "Utilities": [
        "electricity",
        "bescom",
        "msedcl",
        "tneb",
        "water",
        "gas",
        "piped gas",
        "broadband",
        "jio",
        "airtel",
        "vi ",
        "vodafone",
        "bsnl",
        "tata sky",
        "tataplay",
        "dth",
        "recharge",
    ],
    "Healthcare": [
        "apollo",
        "medplus",
        "1mg",
        "netmeds",
        "pharmeasy",
        "hospital",
        "clinic",
        "doctor",
        "medicine",
        "pharmacy",
        "diagnostic",
        "lab test",
        "pathlab",
    ],
    "Education": [
        "udemy",
        "coursera",
        "byjus",
        "unacademy",
        "vedantu",
        "school fee",
        "college fee",
        "tuition",
        "edtech",
    ],
    "Gym & Fitness": [
        "gym",
        "fitness",
        "cult.fit",
        "healthifyme",
        "yoga",
        "crossfit",
        "sports",
    ],
    "Travel": [
        "irctc",
        "makemytrip",
        "goibibo",
        "cleartrip",
        "yatra",
        "ixigo",
        "airbnb",
        "oyo",
        "hotel booking",
        "flight",
        "train ticket",
        "bus ticket",
    ],
    "Subscriptions": [
        "subscription",
        "membership",
        "renewal",
        "annual plan",
        "monthly plan",
        "icloud",
        "google one",
    ],
    "Personal Care": [
        "salon",
        "spa",
        "haircut",
        "beauty",
        "parlour",
        "grooming",
        "manicure",
    ],
    "Housing": [
        "rent",
        "house rent",
        "maintenance",
        "society fee",
        "water charge",
        "property tax",
        "nobroker",
    ],
    "Investments": [
        "mutual fund",
        "zerodha",
        "groww",
        "upstox",
        "angelone",
        "nps",
        "ppf",
        "fd ",
        "sip",
        "demat",
    ],
    "Transfer": [
        "transfer",
        "neft",
        "rtgs",
        "imps",
        "sent to",
        "paid to",
        "self transfer",
    ],
}

# Paytm UPI tag label → our category name
# "__skip__" means skip the transaction (self-transfers)
_PAYTM_TAG_MAP: dict[str, str] = {
    "food": "Food & Dining",
    "groceries": "Groceries",
    "shopping": "Shopping",
    "fuel": "Transportation",
    "money transfer": "Transfer",
    "medical": "Healthcare",
    "entertainment": "Entertainment",
    "services": "Utilities",
    "bill payments": "Utilities",
    "travel": "Travel",
    "miscellaneous": "Other",
    "self transfer": "__skip__",
}

_MONTH_ABBR = {
    "jan": "01",
    "feb": "02",
    "mar": "03",
    "apr": "04",
    "may": "05",
    "jun": "06",
    "jul": "07",
    "aug": "08",
    "sep": "09",
    "oct": "10",
    "nov": "11",
    "dec": "12",
}


class CategoryMatcher:
    """Matches a transaction description to the best-fit category."""

    def __init__(self, user_categories: list[dict]) -> None:
        # Build lookup: lower-case name → {id, name, icon}
        self._by_name: dict[str, dict] = {c["name"].lower(): c for c in user_categories}

    def match(self, description: str) -> tuple[str, str]:
        """Returns (category_id, category_name). Falls back to 'Other'."""
        desc_lower = description.lower()
        for cat_name, keywords in CATEGORY_KEYWORDS.items():
            for kw in keywords:
                if kw in desc_lower:
                    cat = self._by_name.get(cat_name.lower())
                    if cat:
                        return str(cat["id"]), cat["name"]
        other = self._by_name.get("other")
        if other:
            return str(other["id"]), other["name"]
        return "", "Other"


# =============================================================================
# HEADER NORMALISATION
# =============================================================================


def _normalise_header(h: str) -> str:
    """Collapse a CSV/PDF header to a compact lowercase key."""
    h = h.strip().lower()
    h = re.sub(r"\([^)]*\)", "", h)  # strip (₹), (rs), (inr), (amt.)
    h = re.sub(r"[\s./\-_]+", "", h)  # remove whitespace, dots, slashes, dashes
    return h


def detect_format(headers: list[str]) -> str:
    """
    Detect bank/app format from the header row.
    Returns: "phonepe" | "gpay" | "paytm" | "hdfc" | "sbi" | "generic"
    """
    normalised = [_normalise_header(h) for h in headers]
    joined = " ".join(normalised)

    # PhonePe: has "transactionid" and "wallet"
    if "transactionid" in joined and "wallet" in joined:
        return "phonepe"
    # Google Pay: description + status + amount + date
    if "description" in joined and "status" in joined and "amount" in joined:
        return "gpay"
    # Paytm: "transactiondetails" or "passbook" or Paytm Bank with "particulars"
    if "transactiondetails" in joined or "paytm" in joined or "passbook" in joined:
        return "paytm"
    if "particulars" in joined and ("debit" in joined or "credit" in joined):
        return "paytm"
    if "wallet" in joined and ("debit" in joined or "credit" in joined):
        return "paytm"
    # HDFC: narration + withdrawalamt
    if "narration" in joined and "withdrawalamt" in joined:
        return "hdfc"
    # SBI: txndate + refno (or valuedate + description + debit)
    if "txndate" in joined and "refno" in joined:
        return "sbi"
    if "valuedate" in joined and "description" in joined and "debit" in joined:
        return "sbi"
    return "generic"


# =============================================================================
# DATE / AMOUNT / DESCRIPTION HELPERS
# =============================================================================

_DATE_FORMATS = [
    "%d/%m/%Y",
    "%d-%m-%Y",
    "%Y-%m-%d",
    "%d %b %Y",
    "%b %d, %Y",
    "%d %B %Y",
    "%d/%m/%Y %H:%M:%S",
    "%d-%m-%Y %H:%M:%S",
    "%d %b %Y, %I:%M %p",
    "%d %b %Y %I:%M %p",
    "%m/%d/%Y",
    "%Y/%m/%d",
    "%d %b %Y %H:%M",
]


def _parse_date(raw: str) -> str:
    """Parse various date string formats → ISO 'YYYY-MM-DD'. Returns today on failure."""
    raw = raw.strip()
    raw_date_only = re.split(r"\s+\d{1,2}:\d{2}", raw)[0].strip()
    # Also strip trailing comma (seen in some Paytm exports)
    raw_date_only = raw_date_only.rstrip(",").strip()
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(raw_date_only, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    logger.warning("could_not_parse_date", raw=raw)
    return date.today().isoformat()


def _parse_amount(raw: str) -> Decimal | None:
    """Parse amount string → positive Decimal. Returns None if unparseable or zero."""
    cleaned = re.sub(r"[₹,\s]", "", str(raw).strip()).lstrip("-+")
    # Handle "Rs.750" format from Paytm UPI PDFs
    cleaned = re.sub(r"^[Rr][Ss]\.?", "", cleaned).strip()
    try:
        val = Decimal(cleaned)
        return val if val > 0 else None
    except (InvalidOperation, ValueError):
        return None


def _clean_description(desc: str) -> str:
    """Clean up UPI narration strings into readable merchant names."""
    prefixes = [
        r"^UPI[-/]",
        r"^NEFT[-/]",
        r"^IMPS[-/]",
        r"^TO TRANSFER[-/]UPI[-/]\d+[-/]",
        r"^BY TRANSFER[-/]UPI[-/]\d+[-/]",
        r"^UPI/\d+/",
        r"^P2P[-/]",
        r"^POS[-/]",
    ]
    cleaned = desc.strip()
    for prefix in prefixes:
        cleaned = re.sub(prefix, "", cleaned, flags=re.IGNORECASE).strip()
    # Remove UPI VPA (email-style): xxx@yyy
    cleaned = re.sub(r"\s*\S+@\S+", "", cleaned).strip()
    # Remove long numeric refs at end
    cleaned = re.sub(r"\s+\d{8,}$", "", cleaned).strip()
    return cleaned or desc.strip()


# =============================================================================
# FORMAT-SPECIFIC PARSERS  (CSV DictReader rows)
# =============================================================================


def _parse_phonepe(reader, matcher: CategoryMatcher) -> list[ParsedTransaction]:
    """
    PhonePe CSV:
      Date | Transaction Id | Amount (Rs ) | Type | Wallet | Balance | Payment method | Bank Reference No. | Additional Info
    """
    results = []
    for i, row in enumerate(reader):
        try:
            r = {_normalise_header(k): str(v).strip() for k, v in row.items() if k}
            amount = _parse_amount(r.get("amount", "0"))
            if amount is None:
                continue
            tx_type = "CREDIT" if "from" in r.get("type", "").lower() else "DEBIT"
            desc_raw = r.get("additionalinfo", "") or r.get(
                "paymentmethod", "UPI Transfer"
            )
            description = _clean_description(desc_raw)
            date_str = _parse_date(r.get("date", ""))
            ref = r.get("bankreferenceno", r.get("transactionid", ""))
            cat_id, cat_name = matcher.match(description)
            results.append(
                ParsedTransaction(
                    row_index=i,
                    date=date_str,
                    description=description or "PhonePe Transaction",
                    amount=amount,
                    transaction_type=tx_type,
                    reference=ref,
                    suggested_category_name=cat_name,
                    suggested_category_id=cat_id,
                    raw=str(dict(row)),
                )
            )
        except Exception as e:
            logger.warning("phonepe_row_skipped", row=i, error=str(e))
    return results


def _parse_gpay(reader, matcher: CategoryMatcher) -> list[ParsedTransaction]:
    """
    Google Pay CSV:
      Date | Description | Amount (INR) | Status
    """
    results = []
    for i, row in enumerate(reader):
        try:
            r = {_normalise_header(k): str(v).strip() for k, v in row.items() if k}
            if r.get("status", "").lower() not in ("completed", "success", ""):
                continue
            raw_amount_str = r.get("amount", r.get("amountinr", "0"))
            amount = _parse_amount(raw_amount_str)
            if amount is None:
                continue
            # Negative → you paid (DEBIT); positive → received (CREDIT)
            try:
                tx_type = (
                    "CREDIT"
                    if Decimal(re.sub(r"[₹,\s]", "", raw_amount_str).strip()) > 0
                    else "DEBIT"
                )
            except Exception:
                tx_type = "DEBIT"
            description = _clean_description(r.get("description", "GPay Transaction"))
            date_str = _parse_date(r.get("date", ""))
            cat_id, cat_name = matcher.match(description)
            results.append(
                ParsedTransaction(
                    row_index=i,
                    date=date_str,
                    description=description,
                    amount=amount,
                    transaction_type=tx_type,
                    reference="",
                    suggested_category_name=cat_name,
                    suggested_category_id=cat_id,
                    raw=str(dict(row)),
                )
            )
        except Exception as e:
            logger.warning("gpay_row_skipped", row=i, error=str(e))
    return results


def _parse_paytm(reader, matcher: CategoryMatcher) -> list[ParsedTransaction]:
    """
    Paytm Passbook / Paytm Bank CSV:
      Date | Transaction Details | Debit | Credit | Balance
      — or —
      Date | Transaction ID | Transaction Details | Debit | Credit | Balance | Status
      — or (Paytm Bank statement) —
      Date | Particulars | Ref No | Debit | Credit | Balance
    """
    results = []
    for i, row in enumerate(reader):
        try:
            r = {_normalise_header(k): str(v).strip() for k, v in row.items() if k}

            # Find debit/credit columns — Paytm uses several naming conventions
            debit = _parse_amount(
                r.get(
                    "debit",
                    r.get("debitamount", r.get("withdrawalamt", r.get("dr", "0"))),
                )
                or "0"
            )
            credit = _parse_amount(
                r.get(
                    "credit",
                    r.get("creditamount", r.get("depositamt", r.get("cr", "0"))),
                )
                or "0"
            )

            if debit and debit > 0:
                amount, tx_type = debit, "DEBIT"
            elif credit and credit > 0:
                amount, tx_type = credit, "CREDIT"
            else:
                continue

            # Skip failed / pending transactions
            status = r.get("status", r.get("transactionstatus", "")).lower()
            if status and status not in (
                "success",
                "successful",
                "completed",
                "settled",
                "",
            ):
                continue

            desc_raw = (
                r.get("transactiondetails", "")
                or r.get("particulars", "")
                or r.get("narration", "")
                or r.get("description", "Paytm Transaction")
            )
            description = _clean_description(desc_raw)
            date_str = _parse_date(r.get("date", r.get("transactiondate", "")))
            ref = r.get("transactionid", r.get("refno", ""))

            cat_id, cat_name = matcher.match(description)
            results.append(
                ParsedTransaction(
                    row_index=i,
                    date=date_str,
                    description=description or "Paytm Transaction",
                    amount=amount,
                    transaction_type=tx_type,
                    reference=ref,
                    suggested_category_name=cat_name,
                    suggested_category_id=cat_id,
                    raw=str(dict(row)),
                )
            )
        except Exception as e:
            logger.warning("paytm_row_skipped", row=i, error=str(e))
    return results


def _parse_hdfc(reader, matcher: CategoryMatcher) -> list[ParsedTransaction]:
    """
    HDFC Bank statement:
      Date | Narration | Chq./Ref.No. | Value Dt | Withdrawal Amt. | Deposit Amt. | Closing Balance
    """
    results = []
    for i, row in enumerate(reader):
        try:
            r = {_normalise_header(k): str(v).strip() for k, v in row.items() if k}
            withdrawal = _parse_amount(r.get("withdrawalamt", "0") or "0")
            deposit = _parse_amount(r.get("depositamt", "0") or "0")
            if withdrawal and withdrawal > 0:
                amount, tx_type = withdrawal, "DEBIT"
            elif deposit and deposit > 0:
                amount, tx_type = deposit, "CREDIT"
            else:
                continue
            description = _clean_description(r.get("narration", ""))
            date_str = _parse_date(r.get("date", ""))
            ref = r.get("chqrefno", r.get("refno", ""))
            cat_id, cat_name = matcher.match(description)
            results.append(
                ParsedTransaction(
                    row_index=i,
                    date=date_str,
                    description=description or "HDFC Transaction",
                    amount=amount,
                    transaction_type=tx_type,
                    reference=ref,
                    suggested_category_name=cat_name,
                    suggested_category_id=cat_id,
                    raw=str(dict(row)),
                )
            )
        except Exception as e:
            logger.warning("hdfc_row_skipped", row=i, error=str(e))
    return results


def _parse_sbi(reader, matcher: CategoryMatcher) -> list[ParsedTransaction]:
    """
    SBI Bank statement:
      Txn Date | Value Date | Description | Ref No./Cheque No. | Debit | Credit | Balance
    """
    results = []
    for i, row in enumerate(reader):
        try:
            r = {_normalise_header(k): str(v).strip() for k, v in row.items() if k}
            debit = _parse_amount(r.get("debit", "0") or "0")
            credit = _parse_amount(r.get("credit", "0") or "0")
            if debit and debit > 0:
                amount, tx_type = debit, "DEBIT"
            elif credit and credit > 0:
                amount, tx_type = credit, "CREDIT"
            else:
                continue
            desc_raw = r.get("description", r.get("narration", "SBI Transaction"))
            description = _clean_description(desc_raw)
            date_str = _parse_date(r.get("txndate", r.get("date", "")))
            ref = r.get("refnochequeno", r.get("refno", ""))
            cat_id, cat_name = matcher.match(description)
            results.append(
                ParsedTransaction(
                    row_index=i,
                    date=date_str,
                    description=description,
                    amount=amount,
                    transaction_type=tx_type,
                    reference=ref,
                    suggested_category_name=cat_name,
                    suggested_category_id=cat_id,
                    raw=str(dict(row)),
                )
            )
        except Exception as e:
            logger.warning("sbi_row_skipped", row=i, error=str(e))
    return results


def _parse_generic(reader, matcher: CategoryMatcher) -> list[ParsedTransaction]:
    """
    Generic fallback: auto-detect columns by keyword.
    Works for ICICI, Axis, Kotak, Paytm Bank, Axis and most other banks.
    """
    fieldnames = [f.strip().lower() for f in (reader.fieldnames or [])]
    date_col = next((f for f in fieldnames if "date" in f), None)
    desc_col = next(
        (
            f
            for f in fieldnames
            if any(
                k in f
                for k in [
                    "desc",
                    "narr",
                    "particular",
                    "detail",
                    "remark",
                    "memo",
                    "transaction",
                ]
            )
        ),
        None,
    )
    debit_col = next(
        (
            f
            for f in fieldnames
            if any(k in f for k in ["debit", "withdrawal", "dr", "paid", "amount"])
        ),
        None,
    )
    credit_col = next(
        (
            f
            for f in fieldnames
            if any(k in f for k in ["credit", "deposit", "cr", "received"])
        ),
        None,
    )
    ref_col = next(
        (f for f in fieldnames if any(k in f for k in ["ref", "id", "chq", "txn"])),
        None,
    )

    if not (date_col and desc_col and debit_col):
        logger.warning("generic_parser_missing_columns", fieldnames=fieldnames)
        return []

    results = []
    for i, row in enumerate(reader):
        try:
            r = {k.strip().lower(): str(v).strip() for k, v in row.items() if k}
            debit = _parse_amount(r.get(debit_col, "0") or "0")
            credit = (
                _parse_amount(r.get(credit_col or "", "0") or "0")
                if credit_col
                else None
            )
            if debit and debit > 0:
                amount, tx_type = debit, "DEBIT"
            elif credit and credit > 0:
                amount, tx_type = credit, "CREDIT"
            else:
                continue
            description = _clean_description(r.get(desc_col, "Transaction"))
            date_str = _parse_date(r.get(date_col, ""))
            ref = r.get(ref_col or "", "")
            cat_id, cat_name = matcher.match(description)
            results.append(
                ParsedTransaction(
                    row_index=i,
                    date=date_str,
                    description=description,
                    amount=amount,
                    transaction_type=tx_type,
                    reference=ref,
                    suggested_category_name=cat_name,
                    suggested_category_id=cat_id,
                    raw=str(dict(row)),
                )
            )
        except Exception as e:
            logger.warning("generic_row_skipped", row=i, error=str(e))
    return results


PARSER_MAP = {
    "phonepe": _parse_phonepe,
    "gpay": _parse_gpay,
    "paytm": _parse_paytm,
    "hdfc": _parse_hdfc,
    "sbi": _parse_sbi,
    "generic": _parse_generic,
}


# =============================================================================
# SHARED HELPER
# =============================================================================


def _rows_to_transactions(
    rows: list[dict],
    headers: list[str],
    matcher: CategoryMatcher,
    fmt: str,
) -> list[ParsedTransaction]:
    """Convert a list of raw dicts (from CSV, PDF, or Excel) into ParsedTransactions."""

    class _FakeDictReader:
        """Wraps a list[dict] so format parsers can iterate it like csv.DictReader."""

        def __init__(self, rows, headers):
            self.fieldnames = headers
            self._rows = rows

        def __iter__(self):
            return iter(self._rows)

    reader = _FakeDictReader(rows, headers)
    return PARSER_MAP[fmt](reader, matcher)


def _build_preview_from_list(
    transactions: list[ParsedTransaction],
    fmt: str,
    file_type: str,
    skipped: int = 0,
) -> ImportPreview:
    """Build ImportPreview directly from a parsed transaction list."""
    debit_count = sum(1 for t in transactions if t.transaction_type == "DEBIT")
    credit_count = sum(1 for t in transactions if t.transaction_type == "CREDIT")
    return ImportPreview(
        format_detected=fmt,
        file_type=file_type,
        total_rows=len(transactions),
        debit_count=debit_count,
        credit_count=credit_count,
        skipped_count=skipped,
        transactions=transactions,
    )


# =============================================================================
# CSV PARSING
# =============================================================================


def parse_csv(
    file_content: bytes,
    user_categories: list[dict],
) -> ImportPreview:
    """
    Parse a CSV file → ImportPreview.
    Kept for backward-compat; use parse_file() for new code.
    """
    return parse_file(file_content, "file.csv", user_categories)


def _parse_csv_file(file_content: bytes, matcher: CategoryMatcher) -> ImportPreview:
    """Internal: parse CSV bytes → ImportPreview."""
    # Decode — try UTF-8 (with BOM), then latin-1 (some banks use latin-1)
    try:
        text = file_content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = file_content.decode("latin-1")

    lines = [line.rstrip() for line in text.splitlines() if line.strip()]
    if not lines:
        return ImportPreview("unknown", "csv", 0, 0, 0, 0, [])

    # Skip non-CSV account info rows at the top — find the actual header row
    header_idx = 0
    for idx, line in enumerate(lines[:15]):
        parts = line.split(",")
        if len(parts) >= 3 and any(
            kw in line.lower()
            for kw in [
                "date",
                "amount",
                "narration",
                "description",
                "debit",
                "credit",
                "transaction",
            ]
        ):
            header_idx = idx
            break

    csv_text = "\n".join(lines[header_idx:])
    reader = csv.DictReader(io.StringIO(csv_text))

    if not reader.fieldnames:
        return ImportPreview("unknown", "csv", 0, 0, 0, 0, [])

    fmt = detect_format(list(reader.fieldnames))
    logger.info("csv_import_parsing", format=fmt, headers=reader.fieldnames)

    transactions = PARSER_MAP[fmt](reader, matcher)

    # Count skipped rows
    try:
        total_raw = sum(1 for _ in csv.DictReader(io.StringIO(csv_text)))
        skipped = max(0, total_raw - len(transactions))
    except Exception:
        skipped = 0

    return ImportPreview(
        format_detected=fmt,
        file_type="csv",
        total_rows=len(transactions),
        debit_count=sum(1 for t in transactions if t.transaction_type == "DEBIT"),
        credit_count=sum(1 for t in transactions if t.transaction_type == "CREDIT"),
        skipped_count=skipped,
        transactions=transactions,
    )


# =============================================================================
# EXCEL (.xlsx) PARSING
# =============================================================================


def _is_paytm_upi_excel(headers: list[str], sheet_name: str) -> bool:
    """
    Return True if this Excel sheet is a Paytm UPI Statement
    (Passbook Payment History format).

    Detection:
      - Sheet name contains "passbook" or "payment history"  — OR —
      - Headers have "UPI Ref No." + "Transaction Details" + "Your Account"
        (no separate Debit / Credit columns — all rows are debit payments)
    """
    joined = " ".join(_normalise_header(h) for h in headers)
    name_l = sheet_name.lower()
    if any(kw in name_l for kw in ("passbook payment", "passbook", "payment history")):
        return True
    return (
        "upirefno" in joined
        and "transactiondetails" in joined
        and "youraccount" in joined
    )


def _parse_paytm_upi_excel_rows(
    rows: list[dict], matcher: CategoryMatcher
) -> list[ParsedTransaction]:
    """
    Parse rows from a Paytm UPI Statement Excel sheet (Passbook Payment History).

    Column layout:
      Date | Time | Transaction Details | Other Transaction Details |
      Your Account | Amount | UPI Ref No. | Order ID | Remarks | Tags | Comment

    Rules:
      - ALL rows are DEBIT (payments made by the user — this sheet has no credits)
      - Skip rows tagged "# Self Transfer"
      - Tags column: "# Money Transfer", "# Food", "# Groceries" etc.
        (may contain emoji or special chars between # and tag name)
      - Amount format: "₹750.00" — always positive in this format
    """
    results: list[ParsedTransaction] = []
    for i, row in enumerate(rows):
        try:
            r = {_normalise_header(k): str(v).strip() for k, v in row.items() if k}

            # ── Skip self-transfers ──────────────────────────────────────
            tag_raw = r.get("tags", r.get("tag", ""))
            # Strip "#", emojis, special chars → plain lowercase text
            tag_clean = re.sub(r"[^a-z\s]", "", tag_raw.lower()).strip()
            if "self transfer" in tag_clean:
                continue

            # ── Amount ──────────────────────────────────────────────────
            amount = _parse_amount(r.get("amount", "0"))
            if amount is None:
                continue

            # ── Date (DD/MM/YYYY format in Paytm UPI Excel) ──────────────
            date_str = _parse_date(r.get("date", ""))

            # ── Description ──────────────────────────────────────────────
            desc_raw = (
                r.get("transactiondetails", "")
                or r.get("description", "")
                or "Paytm UPI"
            )
            description = _clean_description(desc_raw)

            # ── UPI Reference ────────────────────────────────────────────
            ref = re.sub(r"\s+", "", r.get("upirefno", r.get("upiref", "")))

            # ── Category from Tags column ─────────────────────────────────
            # Extract tag name: strip "# " + any emoji/special chars before letters
            tag_name_match = re.search(r"#\s*[^\w]*\s*(.+)", tag_raw)
            tag_name = tag_name_match.group(1).strip() if tag_name_match else ""
            tag_lower = tag_name.lower()

            cat_id, cat_name = "", "Other"
            for tag_key, mapped_cat in _PAYTM_TAG_MAP.items():
                if tag_key in tag_lower:
                    if mapped_cat == "__skip__":
                        break
                    cat_info = matcher._by_name.get(mapped_cat.lower())
                    if cat_info:
                        cat_id = str(cat_info["id"])
                        cat_name = cat_info["name"]
                    else:
                        cat_name = mapped_cat
                    break
            else:
                cat_id, cat_name = matcher.match(description)

            results.append(
                ParsedTransaction(
                    row_index=i,
                    date=date_str,
                    description=description or "Paytm UPI",
                    amount=amount,
                    transaction_type="DEBIT",
                    reference=ref,
                    suggested_category_name=cat_name,
                    suggested_category_id=cat_id,
                    raw=str(row),
                )
            )
        except Exception as e:
            logger.warning("paytm_upi_excel_row_skipped", row=i, error=str(e))

    logger.info("paytm_upi_excel_parsed", count=len(results))
    return results


def _parse_xlsx_file(file_content: bytes, matcher: CategoryMatcher) -> ImportPreview:
    """
    Parse Excel (.xlsx / .xls) bank statement using openpyxl.

    Sheet selection strategy (important — Paytm UPI xlsx has 2 sheets):
      1. Prefer a sheet whose NAME contains "passbook", "payment history", etc.
      2. Otherwise: first sheet that isn't named "summary" / "overview"
      3. Fallback: active / first sheet

    Format detection:
      - Paytm UPI Excel (Passbook Payment History) → specialized parser
      - All other formats → same detect_format() path as CSV

    Why openpyxl data_only=True?
      Reads computed cell values rather than raw formulas.
      Bank exports may use =SUM() for balance columns — we want the number.
    """
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl_not_installed")
        return ImportPreview("unknown", "xlsx", 0, 0, 0, 0, [])

    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(file_content), data_only=True, read_only=True
        )
    except Exception as e:
        logger.error("xlsx_open_failed", error=str(e))
        return ImportPreview("unknown", "xlsx", 0, 0, 0, 0, [])

    if not wb.sheetnames:
        return ImportPreview("unknown", "xlsx", 0, 0, 0, 0, [])

    # ── Sheet selection ──────────────────────────────────────────────────────
    data_keywords = (
        "passbook payment",
        "payment history",
        "transaction",
        "history",
        "passbook",
    )
    skip_keywords = ("summary", "overview", "cover", "info")

    ws = None
    for name in wb.sheetnames:  # prefer data-named sheet
        if any(kw in name.lower() for kw in data_keywords):
            ws = wb[name]
            break
    if ws is None:
        for name in wb.sheetnames:  # skip summary-named sheets
            if not any(kw in name.lower() for kw in skip_keywords):
                ws = wb[name]
                break
    if ws is None:
        ws = wb.active or wb.worksheets[0]  # last resort

    sheet_name: str = getattr(ws, "title", "") or ""
    all_rows = list(ws.rows)
    if not all_rows:
        return ImportPreview("unknown", "xlsx", 0, 0, 0, 0, [])

    # ── Find header row (skip preamble: bank logo, account info, etc.) ───────
    header_kw = [
        "date",
        "amount",
        "narration",
        "description",
        "debit",
        "credit",
        "transaction",
        "upi",
        "ref",
    ]
    header_row_idx = 0
    headers: list[str] = []

    for idx, row in enumerate(all_rows[:20]):
        values = [
            str(cell.value).strip() if cell.value is not None else "" for cell in row
        ]
        if not any(values):
            continue
        if any(kw in " ".join(values).lower() for kw in header_kw):
            header_row_idx = idx
            headers = values
            break

    if not headers:  # fallback: first non-empty row
        for idx, row in enumerate(all_rows):
            values = [
                str(cell.value).strip() if cell.value is not None else ""
                for cell in row
            ]
            if any(values):
                header_row_idx = idx
                headers = values
                break

    if not headers:
        return ImportPreview("unknown", "xlsx", 0, 0, 0, 0, [])

    while headers and not headers[-1]:  # strip empty trailing cols
        headers.pop()

    n_cols = len(headers)

    # ── Convert data rows to dicts ───────────────────────────────────────────
    data_rows: list[dict] = []
    for row in all_rows[header_row_idx + 1 :]:
        values = [
            str(cell.value).strip() if cell.value is not None else "" for cell in row
        ]
        values = (values + [""] * n_cols)[:n_cols]
        rd = dict(zip(headers, values, strict=False))
        if any(v for v in rd.values()):
            data_rows.append(rd)

    if not data_rows:
        return ImportPreview("unknown", "xlsx", 0, 0, 0, 0, [])

    # ── Route: Paytm UPI Excel vs standard formats ───────────────────────────
    if _is_paytm_upi_excel(headers, sheet_name):
        logger.info("xlsx_routed_paytm_upi", sheet=sheet_name, rows=len(data_rows))
        txns = _parse_paytm_upi_excel_rows(data_rows, matcher)
        return _build_preview_from_list(
            txns,
            "paytm",
            "xlsx",
            skipped=max(0, len(data_rows) - len(txns)),
        )

    fmt = detect_format(headers)
    logger.info("xlsx_import_parsing", format=fmt, headers=headers, sheet=sheet_name)
    txns = _rows_to_transactions(data_rows, headers, matcher, fmt)
    return _build_preview_from_list(
        txns,
        fmt,
        "xlsx",
        skipped=max(0, len(data_rows) - len(txns)),
    )


# =============================================================================
# PDF PARSING  (pdfplumber)
# =============================================================================


def _is_paytm_upi_pdf(first_page_text: str) -> bool:
    """
    Return True if the first-page text belongs to a Paytm UPI Statement PDF.

    Paytm UPI Statement PDFs have a card-layout (no bordered tables) and
    contain specific column headers in the text: "Transaction Details",
    "Notes & Tags".  Normal Paytm Bank PDFs have a standard table structure
    that pdfplumber can extract directly.
    """
    t = first_page_text.lower()
    return (
        ("paytm" in t and "statement" in t and "transaction details" in t)
        or ("notes" in t and "tags" in t and "transaction details" in t)
        or ("paytm upi" in t)
    )


def _parse_paytm_upi_pdf(
    file_content: bytes, matcher: CategoryMatcher
) -> ImportPreview:
    """
    Parse Paytm UPI Statement PDF — card-layout, no bordered tables.

    Layout (5 visual columns per transaction row):
      Date & Time | Transaction Details | Notes & Tags | Your Account | Amount

    Strategy:
      1. pdfplumber.extract_words() → list of {text, x0, x1, top} dicts
      2. Assign each word to one of the 5 columns by x-position
      3. Group words into logical lines by y-coordinate (3pt tolerance)
      4. Group lines into transaction blocks — a block starts on a line
         where col-0 begins with a 1-2 digit day number
      5. Merge all column text within a block, extract fields, create transaction

    Amount format: "- Rs.750"   (DEBIT)
                   "Rs.3,000"   (CREDIT — received / self-transfer positive)
    Self-transfers: tagged "# Self Transfer" → skipped
    """
    import pdfplumber

    transactions: list[ParsedTransaction] = []
    tx_idx = 0

    with pdfplumber.open(io.BytesIO(file_content)) as pdf:
        # ── Detect year from header (e.g., "APR'26" → 2026) ─────────────
        first_text = pdf.pages[0].extract_text() or "" if pdf.pages else ""
        year_match = re.search(r"[''`'](\d{2})(?:\b)", first_text)
        try:
            year = 2000 + int(year_match.group(1)) if year_match else date.today().year
        except Exception:
            year = date.today().year

        for page in pdf.pages:
            page_width = float(page.width or 612)

            # ── Column x-boundaries (fraction of page width) ──────────────
            # These approximate the visual column positions in Paytm UPI PDFs.
            # Adjust if user's PDF has different margins.
            col_x = [
                page_width * 0.00,  # col 0: Date & Time
                page_width * 0.13,  # col 1: Transaction Details
                page_width * 0.53,  # col 2: Notes & Tags
                page_width * 0.70,  # col 3: Your Account
                page_width * 0.85,  # col 4: Amount
                page_width * 1.01,  # sentinel end
            ]

            def word_col(w: dict, _cx: list = col_x) -> int:
                """Return column index (0–4) for a word based on its x-centre."""
                x_mid = (w["x0"] + w["x1"]) / 2
                for ci in range(5):
                    if _cx[ci] <= x_mid < _cx[ci + 1]:
                        return ci
                return 4

            # ── Extract words, skip very-top header text ──────────────────
            words = page.extract_words(x_tolerance=5, y_tolerance=3)
            if not words:
                continue

            # Find y of the column header row ("Date & Time" etc.)
            header_max_y = 0.0
            for w in words:
                if w["text"].lower() in (
                    "date",
                    "time",
                    "details",
                    "tags",
                    "amount",
                    "account",
                ):
                    header_max_y = max(header_max_y, w["top"])

            # ── Group words into y-keyed lines ────────────────────────────
            y_groups: dict[int, list[list[str]]] = {}
            for w in words:
                if w["top"] <= header_max_y + 2:
                    continue  # skip header row words
                y_key = int(round(w["top"] / 3)) * 3
                if y_key not in y_groups:
                    y_groups[y_key] = [[], [], [], [], []]  # 5 cols
                y_groups[y_key][word_col(w)].append(w["text"])

            sorted_lines = sorted(y_groups.items())  # [(y_key, [col0, col1, ...])]

            # ── Identify transaction block starts ──────────────────────────
            # A block starts when col-0 text begins with a 1-2 digit day number.
            def is_tx_start(cols: list[list[str]]) -> bool:
                col0 = " ".join(cols[0]).strip()
                return bool(re.match(r"^\d{1,2}\b", col0))

            # ── Collect blocks ─────────────────────────────────────────────
            blocks: list[list[tuple]] = []
            current: list[tuple] = []
            for y_key, cols in sorted_lines:
                if is_tx_start(cols):
                    if current:
                        blocks.append(current)
                    current = [(y_key, cols)]
                elif current:
                    current.append((y_key, cols))
            if current:
                blocks.append(current)

            # ── Parse each block ───────────────────────────────────────────
            for block in blocks:
                # Merge all words across lines within each column
                merged: list[list[str]] = [[], [], [], [], []]
                for _, cols in block:
                    for ci in range(5):
                        merged[ci].extend(cols[ci])

                col0 = " ".join(merged[0])  # Date & Time: "27 May 11:36 PM"
                col1 = " ".join(merged[1])  # Transaction Details
                col2 = " ".join(merged[2])  # Notes & Tags: "# Money Transfer"
                # merged[3] = Your Account: "Axis Bank - 68" (not used for parsing)
                col4 = " ".join(merged[4])  # Amount: "- Rs.750"

                # ── Skip self-transfers ────────────────────────────────────
                tag_text = col2.strip().lstrip("#").strip().lower()
                if "self transfer" in tag_text:
                    continue

                # ── Parse amount ───────────────────────────────────────────
                amount_str = col4.strip()
                is_debit = "-" in amount_str
                amount = _parse_amount(amount_str)
                if amount is None:
                    continue

                tx_type = "DEBIT" if is_debit else "CREDIT"

                # ── Parse date ─────────────────────────────────────────────
                date_match = re.search(
                    r"(\d{1,2})\s+(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)",
                    col0,
                    re.IGNORECASE,
                )
                if not date_match:
                    # Fallback: try full text of block's first line
                    continue
                day = date_match.group(1).zfill(2)
                month_num = _MONTH_ABBR.get(date_match.group(2).lower()[:3], "01")
                date_str = f"{year}-{month_num}-{day}"

                # ── Extract UPI reference ──────────────────────────────────
                ref_match = re.search(
                    r"UPI\s*Ref\s*(?:No\.?)?[:\s]+(\d+)", col1, re.IGNORECASE
                )
                ref = ref_match.group(1) if ref_match else ""

                # ── Clean description ──────────────────────────────────────
                desc = col1
                desc = re.sub(
                    r"UPI\s*(ID|Ref\s*No\.?)[:\s]+\S+", "", desc, flags=re.IGNORECASE
                )
                desc = re.sub(r"\bon\s*$", "", desc, flags=re.IGNORECASE)
                desc = desc.strip("/ ").strip()
                if not desc:
                    desc = "Paytm UPI Transfer"

                # ── Category from tag, fallback keyword match ──────────────
                cat_id, cat_name = "", "Other"
                for tag_key, mapped_cat in _PAYTM_TAG_MAP.items():
                    if tag_key in tag_text:
                        if mapped_cat == "__skip__":
                            break
                        cat_info = matcher._by_name.get(mapped_cat.lower())
                        if cat_info:
                            cat_id = str(cat_info["id"])
                            cat_name = cat_info["name"]
                        else:
                            cat_name = mapped_cat
                        break
                else:
                    cat_id, cat_name = matcher.match(desc)

                transactions.append(
                    ParsedTransaction(
                        row_index=tx_idx,
                        date=date_str,
                        description=desc,
                        amount=amount,
                        transaction_type=tx_type,
                        reference=ref,
                        suggested_category_name=cat_name,
                        suggested_category_id=cat_id,
                        raw=f"col0={col0!r} col2={col2!r} col4={col4!r}",
                    )
                )
                tx_idx += 1

    logger.info("paytm_upi_pdf_parsed", count=len(transactions), year=year)
    return _build_preview_from_list(transactions, "paytm", "pdf")


def _extract_rows_from_pdf(file_content: bytes) -> tuple[list[str], list[dict]]:
    """
    Extract tabular data from a PDF bank statement using pdfplumber.

    Strategy:
    1. Try page.extract_tables() — works for most bank PDFs with bordered tables
    2. Fall back to page.extract_text() line-by-line parsing for PDFs without borders

    Returns: (headers: list[str], rows: list[dict])
    """
    import pdfplumber

    all_table_rows: list[list[str]] = []
    header_row: list[str] | None = None

    with pdfplumber.open(io.BytesIO(file_content)) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()

            if tables:
                for table in tables:
                    if not table:
                        continue
                    for row in table:
                        # Skip None cells, replace with empty string
                        clean_row = [str(cell).strip() if cell else "" for cell in row]
                        if not any(clean_row):
                            continue  # skip blank rows

                        if header_row is None:
                            # First non-empty row on first page = header
                            if any(re.search(r"[a-zA-Z]", c) for c in clean_row):
                                header_row = clean_row
                                continue
                        else:
                            # Skip repeated headers on subsequent pages
                            if clean_row == header_row:
                                continue
                            all_table_rows.append(clean_row)
            else:
                # Fallback: parse raw text lines (less reliable but works for some PDFs)
                text = page.extract_text() or ""
                lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
                for line in lines:
                    if len(line) < 15:
                        continue
                    parts = re.split(r"\s{2,}|\t", line)
                    if len(parts) < 3:
                        continue
                    if header_row is None and any(
                        re.search(r"[a-zA-Z]", p) for p in parts
                    ):
                        if any(
                            kw in line.lower()
                            for kw in [
                                "date",
                                "narration",
                                "debit",
                                "credit",
                                "amount",
                                "description",
                            ]
                        ):
                            header_row = parts
                            continue
                    elif header_row:
                        all_table_rows.append(parts)

    if not header_row or not all_table_rows:
        return [], []

    # Pad/truncate data rows to match header length
    n_cols = len(header_row)
    normalised_rows = []
    for row in all_table_rows:
        padded = (row + [""] * n_cols)[:n_cols]
        normalised_rows.append(dict(zip(header_row, padded, strict=False)))

    logger.info("pdf_extracted", headers=header_row, row_count=len(normalised_rows))
    return header_row, normalised_rows


# =============================================================================
# UNIFIED ENTRY POINT
# =============================================================================


def parse_file(
    file_content: bytes,
    filename: str,
    user_categories: list[dict],
) -> ImportPreview:
    """
    Main entry point: detects CSV / PDF / Excel, parses, returns ImportPreview.

    Args:
        file_content:    raw bytes of the uploaded file
        filename:        original filename (used to detect extension)
        user_categories: list of {id, name, icon} from CategoryService
    """
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    # Detect file type
    is_pdf = ext == "pdf" or file_content[:4] == b"%PDF"
    is_xlsx = ext in ("xlsx", "xls") and not is_pdf
    # Everything else → CSV (includes .csv and unknown extensions)

    matcher = CategoryMatcher(user_categories)

    if is_pdf:
        return _parse_pdf_file(file_content, matcher)
    elif is_xlsx:
        return _parse_xlsx_file(file_content, matcher)
    else:
        return _parse_csv_file(file_content, matcher)


def _parse_pdf_file(file_content: bytes, matcher: CategoryMatcher) -> ImportPreview:
    """
    Internal: route PDF to the right parser.

    Paytm UPI Statement PDFs use a card-layout without table borders —
    those need the extract_words() approach.  All other bank PDFs (HDFC,
    SBI, Paytm Bank) use bordered tables that pdfplumber can extract directly.
    """
    # ── Detect Paytm UPI card-layout PDF ──────────────────────────────────
    try:
        import pdfplumber

        with pdfplumber.open(io.BytesIO(file_content)) as pdf:
            first_text = (pdf.pages[0].extract_text() or "") if pdf.pages else ""
        if _is_paytm_upi_pdf(first_text):
            logger.info("pdf_routed_to_paytm_upi_parser")
            return _parse_paytm_upi_pdf(file_content, matcher)
    except Exception as e:
        logger.warning("pdf_type_detection_failed", error=str(e))

    # ── Standard table-based PDF ───────────────────────────────────────────
    try:
        headers, rows = _extract_rows_from_pdf(file_content)
    except Exception as e:
        logger.error("pdf_extraction_failed", error=str(e))
        return ImportPreview("unknown", "pdf", 0, 0, 0, 0, [])

    if not headers or not rows:
        return ImportPreview("unknown", "pdf", 0, 0, 0, 0, [])

    fmt = detect_format(headers)
    logger.info("pdf_import_parsing", format=fmt, headers=headers, rows=len(rows))

    transactions = _rows_to_transactions(rows, headers, matcher, fmt)
    return _build_preview_from_list(
        transactions, fmt, "pdf", skipped=max(0, len(rows) - len(transactions))
    )


# legacy alias — kept for any code that imports _build_preview directly
def _build_preview(
    transactions: list[ParsedTransaction],
    fmt: str,
    file_type: str,
    raw_text: str,
) -> ImportPreview:
    """Deprecated: use _build_preview_from_list instead."""
    try:
        total_raw = sum(1 for _ in csv.DictReader(io.StringIO(raw_text)))
        skipped = max(0, total_raw - len(transactions))
    except Exception:
        skipped = 0
    return _build_preview_from_list(transactions, fmt, file_type, skipped)
